import fitz
from openpyxl import load_workbook
import os
import argparse
from tqdm import tqdm   

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="highlight components in pdf")

    # Define arguments
    parser.add_argument("--pdf", required=True, help="Pickle File of stored data ")
    parser.add_argument("--inp", required=True, help="PGA CSV file path")
    parser.add_argument("--o", required=True, help="Refdes CSV file path")

    args = parser.parse_args()

    # Assign to variables
    pdf_path = args.pdf
    inp_path = args.inp
    outfile = args.o
    

    shell= True
    # cwd = os.path.dirname(__file__)
    
    print ('received files ')
    # print (shell)
    # print (f'{pdf_path}\n{inp_path}\n{outfile}\n')

    
    doc = fitz.open(pdf_path)
    logs = []
    highlight_terms = []
    if inp_path.endswith(".xlsx") or inp_path.endswith(".xls"):
        try:
            wb = load_workbook(inp_path)
            sheet = wb.active

            # Validate headers
            headers = [cell.value for cell in sheet[1]]
            if "Level" not in headers or "BOM.Ref Des" not in headers:
                logs.append("Missing headers: 'Level' or 'BOM.Ref Des'")
            else:
                level_idx = headers.index("Level") 
                refdes_idx = headers.index("BOM.Ref Des") 
                

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    level = row[level_idx ]
                    refdes = row[refdes_idx ]
                    if level == None or refdes == None: continue
                    # print (level, refdes, type (level))
                    
                    
                    if int(level) == 1 :
                        # print (level, refdes)
                        # print ()
                        highlight_terms.extend([t.strip() for t in str(refdes).split(",") if t.strip()])
            print("Excel processed successfully.")
        except Exception as e:
            logs.append(f"Error reading Excel: {e}")
            
            
    
    elif inp_path.endswith('txt'):
        with open(inp_path) as f:
            x= f.read()
            print (x)
            highlight_terms = [t.strip() for t in x.split(",") if t.strip()]
    else:
        print ('unknown file given as input. Stopping the processing')
        raise SystemExit()
        
    # print (f'items= {highlight_terms}')
    L = len(highlight_terms)
    print (f'# of items = {L}')
    if shell: pbar = tqdm(total=L)    
    
    
    # Highlight terms
    Found_flag_list = [False]* len(highlight_terms)
    count = 0;
    k = 100 / L
    

    for page in doc:
        text = page.get_text("text")  # Extract page text
        words = text.split()          # Split into words for whole-word check

        for idx, term in enumerate(highlight_terms):
            if not Found_flag_list[idx] and term in words:  # Whole-word match check
                areas = page.search_for(term)  # Basic search
                if areas:
                    # Highlight only the first occurrence
                    annot = page.add_highlight_annot(areas[0])
                    annot.set_colors(stroke=(0, 1, 0))  # Green
                    annot.update()
                    Found_flag_list[idx] = True  # Mark as found
                    count += 1
                    # print (f'{count}/{L}')
                    if shell: pbar.update(1)
                
    # Save processed PDF
    doc.save(outfile)

    # Save logs
    # with open(os.path.join(OUTPUT_DIR, "logs.txt"), "w") as f:
        # f.write("\n".join(logs))
        
    for line in logs:
        print (line)

    # print("Processing complete. Please download now")